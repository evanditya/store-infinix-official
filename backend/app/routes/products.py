from fastapi import APIRouter, Depends, Request, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
import asyncio
import threading
from sqlalchemy.orm import Session, joinedload
from app.database import get_db
from app.models import Product, ProductImage, ProductVariant, gen_id
from app.routes.auth import get_current_user, has_perm, is_staff
import json
import os
import re
import random
import string
import io

router = APIRouter(prefix="/api")

EXCEL_COLUMNS = [
    "Nama Produk", "Harga", "Harga Diskon", "Stok",
    "Berat (gram)", "Panjang (cm)", "Lebar (cm)", "Tinggi (cm)",
    "Kategori", "Deskripsi", "Video Produk", "Varian Produk",
]


def _variants_to_str(variants) -> str:
    """Encode variants as pipe-separated string: Tipe:Nama:Harga:Stok:Tersedia"""
    parts = []
    for v in variants:
        harga = str(int(v.price)) if v.price is not None else ""
        stok = str(v.stock or 0)
        tersedia = "Ya" if v.is_available else "Tidak"
        tipe = v.variant_type or ""
        nama = v.variant_name or ""
        parts.append(f"{tipe}:{nama}:{harga}:{stok}:{tersedia}")
    return " | ".join(parts)


def _str_to_variants(raw: str):
    """Decode pipe-separated variant string back into list of dicts."""
    variants = []
    if not raw or not raw.strip():
        return variants
    for part in raw.split("|"):
        part = part.strip()
        if not part:
            continue
        fields = part.split(":")
        if len(fields) < 2:
            continue
        tipe = fields[0].strip()
        nama = fields[1].strip() if len(fields) > 1 else ""
        harga_raw = fields[2].strip() if len(fields) > 2 else ""
        stok_raw = fields[3].strip() if len(fields) > 3 else "0"
        tersedia_raw = fields[4].strip().lower() if len(fields) > 4 else "ya"
        try:
            harga = float(harga_raw) if harga_raw else None
        except ValueError:
            harga = None
        try:
            stok = int(stok_raw)
        except ValueError:
            stok = 0
        is_available = tersedia_raw not in ("tidak", "no", "false", "0")
        variants.append({
            "variant_type": tipe,
            "variant_name": nama,
            "price": harga,
            "stock": stok,
            "is_available": is_available,
        })
    return variants


SELLER_CONFIG_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "seller_config.json")


def load_seller_config():
    if os.path.exists(SELLER_CONFIG_PATH):
        with open(SELLER_CONFIG_PATH, "r") as f:
            data = json.load(f)
            seller_name = data.get("seller_name", "")
            site_name = data.get("site_name") or seller_name or "Toko Online"
            return {
                "username": data.get("username", ""),
                "seller_name": seller_name,
                "site_name": site_name,
                "profile_picture": data.get("profile_picture", ""),
                "brand_colors": data.get("brand_colors", []),
                "banner": data.get("banner", ""),
                "font": data.get("font", ""),
            }
    return {"username": "seller", "seller_name": "Store", "site_name": "Toko Online", "profile_picture": "", "brand_colors": [], "banner": "", "font": ""}


def product_to_dict(product: Product) -> dict:
    desc_images = []
    if product.description_images:
        try:
            desc_images = json.loads(product.description_images)
        except Exception:
            desc_images = []
    specs = []
    if product.specifications:
        try:
            specs = json.loads(product.specifications)
        except Exception:
            specs = []
    return {
        "id": product.id,
        "name": product.name,
        "slug": product.slug,
        "price": product.price,
        "original_price": product.original_price,
        "category": product.category,
        "description": product.description,
        "description_images": desc_images,
        "specifications": specs,
        "sold_count": product.sold_count,
        "stock": effective_stock(product),
        "rating": product.rating,
        "weight": product.weight or 500,
        "length": product.length or 10,
        "width": product.width or 10,
        "height": product.height or 10,
        "primary_image": resolve_primary_image(product),
        "video_url": product.video_url,
        "images": [{"id": img.id, "image_url": img.image_url, "display_order": img.display_order} for img in product.images],
        "variants": [
            {
                "id": v.id,
                "variant_type": v.variant_type,
                "variant_name": v.variant_name,
                "price": v.price,
                "original_price": v.original_price,
                "price_modifier": v.price_modifier,
                "stock": v.stock,
                "is_available": v.is_available,
            }
            for v in product.variants
        ],
    }


def generate_slug(name: str) -> str:
    base = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    suffix = "".join(random.choices(string.ascii_lowercase + string.digits, k=6))
    return f"{base}-{suffix}"


def effective_stock(product: Product) -> int:
    """Return computed stock.
    For multi-level products the real stock lives in `_combinations` rows
    (Series/Ukuran rows are display-only labels with stock=0). Prefer those
    if present; otherwise sum the single-level variant rows; else product.stock.
    """
    combos = [v for v in product.variants if v.variant_type == "_combinations"]
    if combos:
        return sum(v.stock or 0 for v in combos)
    real_variants = [v for v in product.variants if v.variant_type != "_combinations"]
    if real_variants:
        return sum(v.stock or 0 for v in real_variants)
    return product.stock or 0


def sync_product_stock(product: Product) -> None:
    """Write effective_stock back to product.stock so DB stays in sync."""
    combos = [v for v in product.variants if v.variant_type == "_combinations"]
    if combos:
        product.stock = sum(v.stock or 0 for v in combos)
        return
    real_variants = [v for v in product.variants if v.variant_type != "_combinations"]
    if real_variants:
        product.stock = sum(v.stock or 0 for v in real_variants)


def resolve_primary_image(product: Product) -> str:
    """Return primary_image, falling back to first gallery image if primary is a local uploads path."""
    pi = product.primary_image or ""
    if pi.startswith("/uploads/") or not pi:
        gallery = sorted(product.images, key=lambda i: i.display_order)
        if gallery:
            return gallery[0].image_url
    return pi


def product_to_list_dict(product: Product) -> dict:
    return {
        "name": product.name,
        "slug": product.slug,
        "price": product.price,
        "original_price": product.original_price,
        "category": product.category,
        "sold_count": product.sold_count,
        "stock": effective_stock(product),
        "rating": product.rating,
        "primary_image": resolve_primary_image(product),
        "is_available": product.is_available if product.is_available is not None else True,
        "variants": [
            {
                "variant_type": v.variant_type,
                "variant_name": v.variant_name,
                "price": v.price,
                "original_price": v.original_price,
                "price_modifier": v.price_modifier,
                "stock": v.stock,
                "is_available": v.is_available,
            }
            for v in product.variants
        ],
    }


@router.get("/products")
async def list_products(
    category: str = None,
    search: str = None,
    page: int = 1,
    limit: int = 20,
    include_inactive: bool = False,
    db: Session = Depends(get_db),
):
    from sqlalchemy import case as sa_case
    query = db.query(Product).options(joinedload(Product.variants), joinedload(Product.images))
    if not include_inactive:
        query = query.filter((Product.is_available == True) | (Product.is_available == None))
    if category:
        query = query.filter(Product.category == category)
    if search:
        query = query.filter(Product.name.ilike(f"%{search}%"))
    total = query.distinct().count()
    page = max(1, page)
    limit = max(1, min(limit, 100))
    # Out-of-stock products go to the end; within each group newest first (top-left = most recently added)
    stock_order = sa_case((Product.stock == 0, 1), else_=0)
    products = query.distinct().order_by(stock_order, Product.created_at.desc().nulls_last(), Product.id.desc()).offset((page - 1) * limit).limit(limit).all()
    seller = load_seller_config()
    return {
        "products": [product_to_list_dict(p) for p in products],
        "seller": seller,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": max(1, (total + limit - 1) // limit),
    }


@router.get("/categories")
async def list_categories(db: Session = Depends(get_db)):
    from sqlalchemy import distinct as sql_distinct
    cats = db.query(sql_distinct(Product.category)).filter(Product.category.isnot(None)).all()
    return {"categories": sorted([c[0] for c in cats if c[0]])}


def _get_purchaseable_variants(variants):
    """Return only the variants a buyer actually selects (combinations if present, else all non-combo)."""
    combos = [v for v in variants if v.variant_type == "_combinations"]
    if combos:
        return combos
    return [v for v in variants if v.variant_type != "_combinations"]


def _apply_sheet_style(ws, header_cols, col_widths):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F2937")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, (col_name, width) in enumerate(zip(header_cols, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 25
    return border, Alignment(horizontal="left", vertical="top", wrap_text=True)


_SHOPEE_ROW1_KEYS = [
    "et_title_product_id", "et_title_product_name", "et_title_variation_id",
    "et_title_variation_name", "et_title_parent_sku", "et_title_variation_sku",
    "et_title_variation_price", "ps_gtin_code", "et_title_variation_stock",
    "ps_minimum_purchase_quantity", "ps_maximum_purchase_quantity",
    "ps_maximum_purchase_quantity_start_date", "ps_maximum_purchase_quantity_time_period",
    "ps_maximum_purchase_quantity_end_date", "et_title_reason",
    "", "", "", "", "", "", "", "", "",
]

_SHOPEE_DISPLAY_HEADERS = [
    "Kode Produk", "Nama Produk", "Kode Variasi", "Nama Variasi",
    "SKU Induk", "SKU", "Harga", "GTIN", "Stok",
    "Min. Jumlah Pembelian", "Maks. Jumlah Pembelian",
    "Maks. Jumlah Pembelian - Tanggal Mulai", "Maks. Jumlah Pembelian - Jumlah Hari",
    "Maks. Jumlah Pembelian - Tanggal Berakhir", "Alasan Gagal",
    "Harga Diskon", "Berat (gram)", "Panjang (cm)", "Lebar (cm)", "Tinggi (cm)",
    "Kategori", "Deskripsi", "Video Produk", "Tersedia (Ya/Tidak)",
]

_SHOPEE_ROW4 = [
    "", "", "", "", "", "", "Wajib", "", "Wajib",
    "", "", "", "", "", "",
    "", "", "", "", "", "", "", "", "",
]

_COL_WIDTHS = [
    14, 42, 14, 35, 12, 12, 14, 10, 10,
    10, 10, 10, 10, 10, 12,
    14, 12, 12, 12, 12, 20, 55, 35, 16,
]


@router.get("/products/export-excel")
async def export_products_excel(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not has_perm(user, "products"):
        return JSONResponse({"error": "Akses ditolak"}, status_code=403)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    products = db.query(Product).options(joinedload(Product.variants)).order_by(Product.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── Style helpers ────────────────────────────────────────────────────────
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)

    def _hdr_cell(row, col, value, bg="1F4E78"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = center
        c.border = border
        return c

    def _extra_hdr_cell(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="145A32")
        c.alignment = center
        c.border = border
        return c

    # ── Row 1: Shopee internal keys ──────────────────────────────────────────
    for col_idx, key in enumerate(_SHOPEE_ROW1_KEYS, start=1):
        c = ws.cell(row=1, column=col_idx, value=key)
        c.font = Font(size=9, color="999999")

    # ── Row 2: metadata placeholder ──────────────────────────────────────────
    ws.cell(row=2, column=1, value="sales_info")

    # ── Row 3: display headers ───────────────────────────────────────────────
    for col_idx, hdr in enumerate(_SHOPEE_DISPLAY_HEADERS, start=1):
        if col_idx <= 15:
            _hdr_cell(3, col_idx, hdr)
        else:
            _extra_hdr_cell(3, col_idx, hdr)
        ws.column_dimensions[ws.cell(3, col_idx).column_letter].width = _COL_WIDTHS[col_idx - 1]

    # ── Row 4: "Wajib" markers ───────────────────────────────────────────────
    for col_idx, val in enumerate(_SHOPEE_ROW4, start=1):
        if val:
            c = ws.cell(row=4, column=col_idx, value=val)
            c.font = Font(bold=True, color="C0392B", size=9)
            c.alignment = center

    # ── Hide always-empty columns & mark them ────────────────────────────────
    from openpyxl.utils import get_column_letter
    _EMPTY_COLS = [1, 3, 5, 6, 8, 10, 11, 12, 13, 14, 15]   # A C E F H J-O
    for col_idx in _EMPTY_COLS:
        col_letter = get_column_letter(col_idx)
        cd = ws.column_dimensions[col_letter]
        cd.hidden = True
        cd.width = _COL_WIDTHS[col_idx - 1]
        # put a note in row 4 so it's visible when user unhides
        if not ws.cell(row=4, column=col_idx).value:
            c = ws.cell(row=4, column=col_idx, value="tidak perlu diisi")
            c.font = Font(italic=True, color="AAAAAA", size=8)
            c.alignment = center

    # ── Rows 5-6: empty spacers ──────────────────────────────────────────────
    ws.row_dimensions[1].height = 15
    ws.row_dimensions[2].height = 13
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 15

    # ── Data rows starting at row 7 ──────────────────────────────────────────
    row_idx = 7
    for p in products:
        display_variants = _get_purchaseable_variants(p.variants)

        def _data_row(variant_name, price, diskon, stock, tersedia):
            nonlocal row_idx
            vals = [
                "",                                         # A: Kode Produk (empty — no Shopee ID)
                p.name,                                     # B: Nama Produk
                "",                                         # C: Kode Variasi (empty)
                variant_name or "",                         # D: Nama Variasi
                "", "",                                     # E-F: SKU Induk, SKU
                price,                                      # G: Harga
                "",                                         # H: GTIN
                stock,                                      # I: Stok
                "", "", "", "", "",                         # J-N: pembelian limits (empty)
                "",                                         # O: Alasan Gagal
                diskon or "",                               # P: Harga Diskon
                p.weight or 500,                            # Q: Berat
                p.length or 10,                             # R: Panjang
                p.width or 10,                              # S: Lebar
                p.height or 10,                             # T: Tinggi
                p.category or "",                           # U: Kategori
                p.description or "",                        # V: Deskripsi
                p.video_url or "",                          # W: Video
                tersedia,                                   # X: Tersedia
            ]
            for col_idx2, value in enumerate(vals, start=1):
                c = ws.cell(row=row_idx, column=col_idx2, value=value)
                c.border = border
                c.alignment = left
                if col_idx2 in (7, 16):
                    c.number_format = "#,##0"
            ws.row_dimensions[row_idx].height = 16
            row_idx += 1

        if display_variants:
            for v in display_variants:
                vname = (v.variant_name or "").replace(" / ", ",") if (v.variant_type == "_combinations" or " / " in (v.variant_type or "")) else (v.variant_name or "")
                _data_row(
                    variant_name=vname,
                    price=int(v.price) if v.price is not None else int(p.price or 0),
                    diskon=int(v.original_price) if v.original_price else None,
                    stock=v.stock if v.stock is not None else "",
                    tersedia="Ya" if v.is_available else "Tidak",
                )
        else:
            _data_row(
                variant_name="",
                price=int(p.price or 0),
                diskon=int(p.original_price) if p.original_price else None,
                stock=p.stock or 0,
                tersedia="Ya" if (p.is_available is not False) else "Tidak",
            )

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=produk-shopee.xlsx"},
    )


@router.post("/products/import-excel")
async def import_products_excel(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not has_perm(user, "products"):
        return JSONResponse({"error": "Akses ditolak"}, status_code=403)

    if not file.filename.endswith((".xlsx", ".xls")):
        return JSONResponse({"error": "File harus berformat .xlsx"}, status_code=400)

    from openpyxl import load_workbook

    content = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    except Exception as e:
        return JSONResponse({"error": f"File Excel tidak valid: {e}"}, status_code=400)

    # ── Pre-load ALL products + variants into memory (one query) ────────────
    all_products = db.query(Product).options(joinedload(Product.variants)).all()
    prod_by_name = {p.name: p for p in all_products}
    # variant lookup: (product_id, variant_name) -> variant object
    var_by_key = {(v.product_id, v.variant_name): v
                  for p in all_products for v in p.variants}

    def _get_col(header_row):
        return {str(h).strip() if h is not None else "": idx
                for idx, h in enumerate(header_row)}

    def _cell(row, col_map, name):
        idx = col_map.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    def _parse_num(val):
        if val is None:
            return None
        try:
            return float(str(val).replace(",", "").replace(".", "").strip()
                         if isinstance(val, str) else val)
        except (ValueError, TypeError):
            return None

    def _num_eq(old, new):
        """True if old and new represent the same numeric value (ignores float/int mismatch)."""
        if old is None and new is None:
            return True
        if old is None or new is None:
            return False
        try:
            return round(float(old)) == round(float(new))
        except (TypeError, ValueError):
            return False

    total_prod = updated_prod = skipped_prod = 0
    total_var  = updated_var  = skipped_var  = 0
    not_found: list = []
    errors: list = []

    # ── Sheet 1: Product info (no DB queries inside loop) ───────────────────
    ws1 = wb.active
    rows1 = list(ws1.iter_rows(values_only=True))
    if rows1:
        col1 = _get_col(rows1[0])
        if "Nama Produk" not in col1:
            return JSONResponse(
                {"error": "Sheet 'Produk': kolom 'Nama Produk' tidak ditemukan"},
                status_code=400)

        for row_num, row in enumerate(rows1[1:], start=2):
            nama = str(_cell(row, col1, "Nama Produk") or "").strip()
            if not nama:
                continue
            total_prod += 1
            product = prod_by_name.get(nama)
            if not product:
                not_found.append(nama)
                continue
            try:
                changed = False

                harga = _parse_num(_cell(row, col1, "Harga"))
                if harga is not None and not _num_eq(product.price, harga):
                    product.price = harga; changed = True

                # Support both "Harga Diskon" (new) and "Harga Coret" (old column name)
                hc_col = "Harga Diskon" if "Harga Diskon" in col1 else "Harga Coret"
                hc_raw = _cell(row, col1, hc_col)
                if hc_raw is not None and str(hc_raw).strip() != "":
                    hc = _parse_num(hc_raw)
                    new_hc = hc if hc and hc > 0 else None
                    if not _num_eq(product.original_price, new_hc):
                        product.original_price = new_hc; changed = True
                elif hc_raw is not None and str(hc_raw).strip() == "" and product.original_price is not None:
                    product.original_price = None; changed = True

                stok_raw = _cell(row, col1, "Stok (tanpa varian)")
                if stok_raw is not None and str(stok_raw).strip() not in ("", "-"):
                    try:
                        new_stok = int(float(str(stok_raw).strip()))
                        if product.stock != new_stok:
                            product.stock = new_stok; changed = True
                    except (ValueError, TypeError):
                        pass

                for attr, cname in [("weight", "Berat (gram)"), ("length", "Panjang (cm)"),
                                     ("width", "Lebar (cm)"), ("height", "Tinggi (cm)")]:
                    v = _cell(row, col1, cname)
                    if v is not None and str(v).strip():
                        try:
                            new_v = int(float(str(v).strip()))
                            if getattr(product, attr) != new_v:
                                setattr(product, attr, new_v); changed = True
                        except (ValueError, TypeError):
                            pass

                for attr, cname in [("category", "Kategori"), ("description", "Deskripsi"),
                                     ("video_url", "Video Produk")]:
                    v = _cell(row, col1, cname)
                    if v is not None:
                        new_v = str(v).strip() or None
                        if getattr(product, attr) != new_v:
                            setattr(product, attr, new_v); changed = True

                if changed:
                    updated_prod += 1
                else:
                    skipped_prod += 1
            except Exception as e:
                errors.append({"row": row_num, "name": nama, "error": str(e)})

    # ── Sheet 2: Variants (no DB queries inside loop) ────────────────────────
    ws2 = wb["Varian"] if "Varian" in wb.sheetnames else None
    affected_product_ids: set = set()
    if ws2 is not None:
        rows2 = list(ws2.iter_rows(values_only=True))
        if len(rows2) > 1:
            col2 = _get_col(rows2[0])
            for row_num, row in enumerate(rows2[1:], start=2):
                nama = str(_cell(row, col2, "Nama Produk") or "").strip()
                nama_varian = str(_cell(row, col2, "Nama Varian") or "").strip()
                if not nama or not nama_varian:
                    continue
                total_var += 1
                product = prod_by_name.get(nama)
                if not product:
                    if nama not in not_found:
                        not_found.append(nama)
                    continue
                variant = var_by_key.get((product.id, nama_varian))
                if not variant:
                    errors.append({"row": row_num,
                                   "name": f"{nama} → {nama_varian}",
                                   "error": "Nama varian tidak cocok"})
                    continue
                try:
                    changed = False

                    # Support both "Harga Asli" (new) and "Harga" (old column name)
                    harga_raw = _cell(row, col2, "Harga Asli") if "Harga Asli" in col2 else _cell(row, col2, "Harga")
                    harga = _parse_num(harga_raw)
                    if harga is not None and not _num_eq(variant.price, harga):
                        variant.price = harga; changed = True

                    hd_raw = _cell(row, col2, "Harga Diskon")
                    if hd_raw is not None and str(hd_raw).strip() != "":
                        hd = _parse_num(hd_raw)
                        new_hd = hd if hd and hd > 0 else None
                        if not _num_eq(variant.original_price, new_hd):
                            variant.original_price = new_hd; changed = True
                    elif hd_raw is not None and str(hd_raw).strip() == "" and variant.original_price is not None:
                        variant.original_price = None; changed = True

                    stok_raw = _cell(row, col2, "Stok")
                    if stok_raw is not None and str(stok_raw).strip():
                        try:
                            new_stok = int(float(str(stok_raw).strip()))
                            if variant.stock != new_stok:
                                variant.stock = new_stok; changed = True
                        except (ValueError, TypeError):
                            pass

                    tersedia = str(_cell(row, col2, "Tersedia (Ya/Tidak)") or "").strip().lower()
                    if tersedia:
                        new_avail = tersedia not in ("tidak", "no", "false", "0")
                        if variant.is_available != new_avail:
                            variant.is_available = new_avail; changed = True

                    if changed:
                        updated_var += 1
                        affected_product_ids.add(product.id)
                    else:
                        skipped_var += 1
                except Exception as e:
                    errors.append({"row": row_num,
                                   "name": f"{nama} → {nama_varian}",
                                   "error": str(e)})

    # ── Sync product.stock from variant stocks for all affected products ──────
    for pid in affected_product_ids:
        prod = next((p for p in prod_by_name.values() if p.id == pid), None)
        if prod:
            sync_product_stock(prod)

    # ── Single commit for everything ─────────────────────────────────────────
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return JSONResponse({"error": f"Gagal menyimpan ke database: {e}"}, status_code=500)

    return {
        "total": total_prod + total_var,
        "updated": updated_prod + updated_var,
        "skipped": skipped_prod + skipped_var,
        "detail": {
            "produk_diperbarui": updated_prod,
            "produk_tidak_berubah": skipped_prod,
            "varian_diperbarui": updated_var,
            "varian_tidak_berubah": skipped_var,
        },
        "not_found": not_found,
        "not_found_count": len(not_found),
        "errors": errors,
        "error_count": len(errors),
    }


@router.get("/products/{slug}")
async def get_product(slug: str, db: Session = Depends(get_db)):
    product = db.query(Product).options(
        joinedload(Product.images),
        joinedload(Product.variants)
    ).filter(Product.slug == slug).first()
    if not product:
        return JSONResponse({"error": "Produk tidak ditemukan"}, status_code=404)
    return {"product": product_to_dict(product)}


@router.post("/products")
async def create_product(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not has_perm(user, "products"):
        return JSONResponse({"error": "Akses ditolak"}, status_code=403)
    body = await request.json()
    slug = generate_slug(body.get("name", "product"))
    product = Product(
        id=gen_id(),
        name=body.get("name", ""),
        slug=slug,
        price=body.get("price", 0),
        original_price=body.get("original_price"),
        category=body.get("category"),
        description=body.get("description"),
        sold_count=body.get("sold_count", 0),
        stock=body.get("stock", 0),
        rating=body.get("rating", 0),
        weight=body.get("weight", 500),
        length=body.get("length", 10),
        width=body.get("width", 10),
        height=body.get("height", 10),
        primary_image=body.get("primary_image"),
        video_url=body.get("video_url"),
    )
    db.add(product)
    for img in body.get("images", []):
        db.add(ProductImage(id=gen_id(), product_id=product.id, image_url=img.get("image_url", img if isinstance(img, str) else ""), display_order=img.get("display_order", 0) if isinstance(img, dict) else 0))
    for idx, v in enumerate(body.get("variants", [])):
        db.add(ProductVariant(
            id=gen_id(), product_id=product.id,
            variant_type=v.get("variant_type"), variant_name=v.get("variant_name", ""),
            price=v.get("price"), original_price=v.get("original_price"),
            price_modifier=v.get("price_modifier", 0),
            stock=v.get("stock", 0), is_available=v.get("is_available", True),
            display_order=idx,
        ))
    all_vars = body.get("variants", []) or []
    combo_vars = [v for v in all_vars if v.get("variant_type") == "_combinations"]
    real_vars = [v for v in all_vars if v.get("variant_type") != "_combinations"]
    if combo_vars:
        product.stock = sum(int(v.get("stock", 0) or 0) for v in combo_vars)
    elif real_vars:
        product.stock = sum(int(v.get("stock", 0) or 0) for v in real_vars)
    db.commit()
    db.refresh(product)
    return {"product": product_to_dict(product)}


@router.put("/products/{slug}")
async def update_product(slug: str, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not has_perm(user, "products"):
        return JSONResponse({"error": "Akses ditolak"}, status_code=403)
    product = db.query(Product).filter(Product.slug == slug).first()
    if not product:
        return JSONResponse({"error": "Produk tidak ditemukan"}, status_code=404)
    body = await request.json()
    for field in ["name", "price", "original_price", "category", "description", "stock", "rating", "weight", "length", "width", "height", "primary_image", "video_url", "sold_count", "is_available"]:
        if field in body:
            setattr(product, field, body[field])

    if "variants" in body:
        db.query(ProductVariant).filter(ProductVariant.product_id == product.id).delete()
        for idx, v in enumerate(body["variants"]):
            db.add(ProductVariant(
                id=gen_id(), product_id=product.id,
                variant_type=v.get("variant_type", ""),
                variant_name=v.get("variant_name", ""),
                price=v.get("price"),
                original_price=v.get("original_price"),
                price_modifier=v.get("price_modifier", 0),
                stock=v.get("stock", 0),
                is_available=v.get("is_available", True),
                display_order=idx,
            ))
        all_vars = body["variants"] or []
        combo_vars = [v for v in all_vars if v.get("variant_type") == "_combinations"]
        real_vars = [v for v in all_vars if v.get("variant_type") != "_combinations"]
        if combo_vars:
            product.stock = sum(int(v.get("stock", 0) or 0) for v in combo_vars)
        elif real_vars:
            product.stock = sum(int(v.get("stock", 0) or 0) for v in real_vars)

    db.commit()
    db.refresh(product)
    return {"product": product_to_dict(product)}


@router.get("/categories")
async def list_categories(db: Session = Depends(get_db)):
    products = db.query(Product.category).distinct().all()
    categories = sorted([p[0] for p in products if p[0]])
    return {"categories": categories}


@router.delete("/products/{slug}")
async def delete_product(slug: str, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not has_perm(user, "products"):
        return JSONResponse({"error": "Akses ditolak"}, status_code=403)
    product = db.query(Product).filter(Product.slug == slug).first()
    if not product:
        return JSONResponse({"error": "Produk tidak ditemukan"}, status_code=404)
    db.delete(product)
    db.commit()
    return {"success": True}


@router.post("/products/repair-stock-json")
async def repair_stock_from_json(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """One-time repair: accept products.json and set product.stock from the
    JSON 'stock' field for any product without real variants. Matches by
    shopee_url id (preferred) then by exact name."""
    user = get_current_user(request, db)
    if not has_perm(user, "products"):
        return JSONResponse({"error": "Akses ditolak"}, status_code=403)
    raw = await file.read()
    try:
        data = json.loads(raw)
    except Exception:
        return JSONResponse({"error": "JSON tidak valid"}, status_code=400)
    products_list = data.get("products", []) if isinstance(data, dict) else (data or [])
    # Bulk-load: avoid N×3 round-trips to the DB
    all_prods = db.query(Product.id, Product.name, Product.shopee_url, Product.stock).all()
    by_shopee = {p.shopee_url: p for p in all_prods if p.shopee_url}
    by_name = {p.name: p for p in all_prods}
    products_with_variants = {pid for (pid,) in db.query(ProductVariant.product_id).distinct().all()}
    updated = 0
    not_found = 0
    skipped_has_variants = 0
    pending: dict = {}  # product_id -> new_stock
    for p in products_list:
        name = (p.get("name") or "").strip()
        if not name:
            continue
        product_url = p.get("product_url", "") or ""
        m = re.search(r"(i\.\d+\.\d+)", product_url)
        shopee_id = m.group(1) if m else ""
        row = by_shopee.get(shopee_id) if shopee_id else None
        if row is None:
            row = by_name.get(name)
        if row is None:
            not_found += 1
            continue
        if row.id in products_with_variants:
            skipped_has_variants += 1
            continue
        try:
            stock_val = int(p.get("stock", 0) or 0)
        except Exception:
            stock_val = 0
        if (row.stock or 0) != stock_val:
            pending[row.id] = stock_val
    # Apply updates in a single bulk operation
    if pending:
        for pid, new_stock in pending.items():
            db.query(Product).filter(Product.id == pid).update({"stock": new_stock}, synchronize_session=False)
            updated += 1
        db.commit()
    return {"success": True, "updated": updated, "not_found": not_found, "skipped_has_variants": skipped_has_variants}


@router.post("/products/recompute-stock")
async def recompute_all_product_stock(request: Request, db: Session = Depends(get_db)):
    """Recompute product.stock from sum of variant.stock for every product.
    Used to repair data after the sync-zip variant-stock bug."""
    user = get_current_user(request, db)
    if not has_perm(user, "products"):
        return JSONResponse({"error": "Akses ditolak"}, status_code=403)
    from sqlalchemy import func
    # Sum stock from _combinations rows (multi-level products) — these hold the real stock.
    combo_rows = (
        db.query(ProductVariant.product_id, func.sum(ProductVariant.stock))
        .filter(ProductVariant.variant_type == "_combinations")
        .group_by(ProductVariant.product_id)
        .all()
    )
    combo_sums = {pid: int(s or 0) for pid, s in combo_rows}
    # Fallback: sum non-combination rows (single-level variant products).
    flat_rows = (
        db.query(ProductVariant.product_id, func.sum(ProductVariant.stock))
        .filter(ProductVariant.variant_type != "_combinations")
        .group_by(ProductVariant.product_id)
        .all()
    )
    flat_sums = {pid: int(s or 0) for pid, s in flat_rows}
    updated = 0
    for prod in db.query(Product).all():
        if prod.id in combo_sums:
            new_stock = combo_sums[prod.id]
        elif prod.id in flat_sums:
            new_stock = flat_sums[prod.id]
        else:
            continue
        if (prod.stock or 0) != new_stock:
            prod.stock = new_stock
            updated += 1
    db.commit()
    return {"success": True, "updated": updated,
            "multi_level": len(combo_sums), "single_level_only": len(flat_sums) - len(combo_sums)}


@router.post("/products/sync-zip")
async def sync_products_zip(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not has_perm(user, "products"):
        return JSONResponse({"error": "Akses ditolak"}, status_code=403)

    import zipfile as zf_mod
    from app.database import SessionLocal

    content = await file.read()
    try:
        zfile = zf_mod.ZipFile(io.BytesIO(content))
    except Exception:
        return JSONResponse({"error": "File bukan zip yang valid"}, status_code=400)

    if "products.json" not in zfile.namelist():
        return JSONResponse({"error": "products.json tidak ditemukan dalam zip"}, status_code=400)

    try:
        raw = json.loads(zfile.read("products.json"))
    except Exception:
        return JSONResponse({"error": "products.json tidak bisa dibaca"}, status_code=400)

    products_list = raw.get("products", [])
    uploads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads")
    os.makedirs(uploads_dir, exist_ok=True)
    zip_names = set(zfile.namelist())
    total = len(products_list)

    loop = asyncio.get_event_loop()
    queue: asyncio.Queue = asyncio.Queue()

    def _parse_price(val: str) -> float:
        # Handle ranges like "Rp5.500.000 - Rp9.790.000" → take lower bound.
        s = str(val)
        if "-" in s:
            s = s.split("-")[0]
        cleaned = re.sub(r"[^\d]", "", s)
        return float(cleaned) if cleaned else 0.0

    def _parse_sold(val: str) -> int:
        m = re.search(r"[\d.]+", str(val))
        return int(m.group().replace(".", "")) if m else 0

    def worker():
        db_w = SessionLocal()
        created = updated = 0
        errors = []
        from datetime import datetime as _dt, timedelta as _td
        sync_base_time = _dt.utcnow()
        try:
            for idx, p in enumerate(products_list):
                try:
                    name = p.get("name", "").strip()
                    if not name:
                        continue

                    product_url = p.get("product_url", "")
                    m = re.search(r"(i\.\d+\.\d+)", product_url)
                    shopee_id = m.group(1) if m else ""

                    price = _parse_price(p.get("price", "0"))
                    sold = _parse_sold(p.get("sold", "0"))
                    stock = int(p.get("stock", 0) or 0)
                    category_raw = p.get("category", "") or ""
                    category = category_raw.split(">")[-1].strip() if category_raw else None

                    product = None
                    if shopee_id:
                        product = db_w.query(Product).filter(Product.shopee_url == shopee_id).first()
                    if product is None:
                        product = db_w.query(Product).filter(Product.name == name).first()

                    is_new = product is None
                    if is_new:
                        product = Product(id=gen_id())
                        base_slug = generate_slug(name)
                        while db_w.query(Product).filter(Product.slug == base_slug).first():
                            base_slug = generate_slug(name)
                        product.slug = base_slug
                        product.weight = 500
                        product.length = product.width = product.height = 10
                        # Preserve Shopee order: sort is newest-first, so the first
                        # product in the file needs the largest timestamp (total - idx)
                        product.created_at = sync_base_time + _td(milliseconds=(len(products_list) - idx))

                    product.name = name
                    product.price = price
                    product.description = p.get("description", "") or ""
                    product.category = category
                    product.sold_count = sold
                    product.shopee_url = shopee_id or (product.shopee_url if not is_new else "")

                    video_list = p.get("all_videos", []) or []
                    if video_list:
                        fv = video_list[0]
                        product.video_url = fv if isinstance(fv, str) else fv.get("url", "")

                    thumb_key = f"images/product_{idx}_thumb.jpg"
                    if thumb_key in zip_names:
                        safe_id = re.sub(r"[^a-z0-9]", "_", shopee_id.lower()) if shopee_id else str(idx)
                        thumb_fn = f"sync_{safe_id}_thumb.jpg"
                        with open(os.path.join(uploads_dir, thumb_fn), "wb") as fout:
                            fout.write(zfile.read(thumb_key))
                        product.primary_image = f"/uploads/{thumb_fn}"

                    if is_new:
                        db_w.add(product)
                        db_w.flush()

                    for v in list(product.variants):
                        db_w.delete(v)
                    db_w.flush()

                    variants_data = p.get("variants", []) or []
                    # Treat variants as absent when no group actually has options
                    has_real_variants = any((vg.get("options") or []) for vg in variants_data)
                    if has_real_variants:
                        non_combo_total = 0
                        combo_total = 0
                        has_combo = False
                        for vgroup in variants_data:
                            vtype = vgroup.get("type", "Pilihan")
                            is_combo = (vtype == "_combinations")
                            if is_combo:
                                has_combo = True
                            for opt in vgroup.get("options", []):
                                v_price = _parse_price(opt.get("price", "0")) or price
                                v_stock = int(opt.get("stock", 0) or 0)
                                if is_combo:
                                    combo_total += v_stock
                                else:
                                    non_combo_total += v_stock
                                db_w.add(ProductVariant(
                                    id=gen_id(), product_id=product.id,
                                    variant_type=vtype, variant_name=opt.get("name", ""),
                                    price=v_price, stock=v_stock,
                                    is_available=bool(opt.get("available", True)),
                                ))
                        db_w.flush()
                        # Multi-level products carry real stock on _combinations rows;
                        # the Series/Ukuran display rows have stock=0 and must be ignored.
                        product.stock = combo_total if has_combo else non_combo_total
                    else:
                        product.stock = stock

                    for img in list(product.images):
                        db_w.delete(img)
                    db_w.flush()

                    for img_idx, img in enumerate(p.get("all_images", []) or []):
                        img_url = img.get("url", "") if isinstance(img, dict) else str(img)
                        if img_url:
                            db_w.add(ProductImage(
                                id=gen_id(), product_id=product.id,
                                image_url=img_url, display_order=img_idx,
                            ))

                    if is_new:
                        created += 1
                    else:
                        updated += 1

                    if (idx + 1) % 20 == 0:
                        db_w.commit()

                except Exception as exc:
                    errors.append(f"{p.get('name','?')[:40]}: {str(exc)[:80]}")

                if (idx + 1) % 5 == 0 or idx == total - 1:
                    loop.call_soon_threadsafe(queue.put_nowait, {
                        "processed": idx + 1, "total": total,
                        "created": created, "updated": updated,
                        "done": False,
                    })

            db_w.commit()
            loop.call_soon_threadsafe(queue.put_nowait, {
                "processed": total, "total": total,
                "created": created, "updated": updated,
                "skipped_errors": len(errors), "errors": errors[:10],
                "done": True, "success": True,
            })
        except Exception as exc:
            db_w.rollback()
            loop.call_soon_threadsafe(queue.put_nowait, {"error": str(exc), "done": True})
        finally:
            db_w.close()

    threading.Thread(target=worker, daemon=True).start()

    async def generate():
        while True:
            item = await queue.get()
            yield f"data: {json.dumps(item)}\n\n"
            if item.get("done"):
                break

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"})
